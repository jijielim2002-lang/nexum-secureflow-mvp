"""
Nexum SecureFlow — Live Data Exporter
Run: python export_live_data.py
Output: nexum-live-data-export.xlsx (same folder)

Requirements: pip install supabase openpyxl
"""
import json, os, sys
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from supabase import create_client

# ── Load credentials from .env.local ─────────────────────────────────────────
env_path = Path(__file__).parent / ".env.local"
if not env_path.exists():
    print("ERROR: .env.local not found next to this script.")
    sys.exit(1)

env = {}
for line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
    line = line.strip()
    if "=" in line and not line.startswith("#"):
        k, _, v = line.partition("=")
        env[k.strip()] = v.strip()

SUPABASE_URL = env.get("NEXT_PUBLIC_SUPABASE_URL", "")
SUPABASE_KEY = env.get("SUPABASE_SERVICE_ROLE_KEY", "")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("ERROR: Missing NEXT_PUBLIC_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY in .env.local")
    sys.exit(1)

sb = create_client(SUPABASE_URL, SUPABASE_KEY)

# ── Helpers ───────────────────────────────────────────────────────────────────
def fetch(table, select="*", order=None, limit=1000):
    q = sb.table(table).select(select).limit(limit)
    if order:
        q = q.order(order, desc=True)
    try:
        r = q.execute()
        return r.data or []
    except Exception as e:
        print(f"  WARN: Could not fetch {table}: {e}")
        return []

def thin():
    s = Side(style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, cols, bg="0369A1"):
    fill = PatternFill("solid", fgColor=bg)
    for i, h in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin()
    ws.row_dimensions[row].height = 30

def cell(ws, row, col, val, bold=False, color="1E293B", bg=None, align="left", wrap=False):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name="Arial", bold=bold, color=color, size=10)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.border = thin()

def title_banner(ws, text, span, bg="0B1929"):
    ws.merge_cells(f"A1:{get_column_letter(span)}1")
    c = ws.cell(row=1, column=1, value=text)
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=14)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

def widths(ws, wlist):
    for i, w in enumerate(wlist, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def row_bg(i):
    return "F8FAFC" if i % 2 == 0 else "FFFFFF"

def fmt(val):
    if val is None:
        return "—"
    if isinstance(val, bool):
        return "Yes" if val else "No"
    if isinstance(val, dict):
        return json.dumps(val)[:120]
    return str(val)

def money(val):
    if val is None:
        return "—"
    try:
        return f"{float(val):,.2f}"
    except:
        return str(val)

# ── Fetch all tables ──────────────────────────────────────────────────────────
print("Connecting to Supabase...")
print(f"URL: {SUPABASE_URL[:40]}...")

print("Fetching companies...")
companies    = fetch("companies",    order="created_at")
print(f"  → {len(companies)} rows")

print("Fetching profiles (users)...")
profiles     = fetch("profiles",     order="created_at")
print(f"  → {len(profiles)} rows")

print("Fetching secured_jobs...")
jobs         = fetch("secured_jobs", order="created_at")
print(f"  → {len(jobs)} rows")

print("Fetching vendor_credit_terms...")
vendor_terms = fetch("vendor_credit_terms", order="created_at")
print(f"  → {len(vendor_terms)} rows")

print("Fetching tradeflow_requests...")
tradeflows   = fetch("tradeflow_requests", order="created_at")
print(f"  → {len(tradeflows)} rows")

print("Fetching trade_chains...")
chains       = fetch("trade_chains", order="created_at")
print(f"  → {len(chains)} rows")

print("Fetching tradecycle_wallets...")
wallets      = fetch("tradecycle_wallets", order="created_at")
print(f"  → {len(wallets)} rows")

print("Fetching company_intelligence_scores...")
scores       = fetch("company_intelligence_scores", order="calculated_at")
print(f"  → {len(scores)} rows")

print("Fetching rfq_requests...")
rfqs         = fetch("rfq_requests", order="created_at")
print(f"  → {len(rfqs)} rows")

print("Fetching rfq_quotes...")
quotes       = fetch("rfq_quotes",   order="created_at")
print(f"  → {len(quotes)} rows")

print("Fetching service_listings...")
listings     = fetch("service_listings", order="created_at")
print(f"  → {len(listings)} rows")

print("Fetching shipment_bundles...")
bundles      = fetch("shipment_bundles", order="created_at")
print(f"  → {len(bundles)} rows")

print("Fetching risk_signals...")
risk_signals = fetch("risk_signals", order="created_at")
print(f"  → {len(risk_signals)} rows")

print("Fetching tradecycle_wallet_transactions...")
wallet_txns  = fetch("tradecycle_wallet_transactions", order="created_at")
print(f"  → {len(wallet_txns)} rows")

# Company lookup: id → name
co_lookup = {c["id"]: c.get("name", c["id"][:8]) for c in companies}

# ── Build workbook ────────────────────────────────────────────────────────────
print("\nBuilding Excel workbook...")
wb = openpyxl.Workbook()

# ── SHEET: Companies ─────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Companies"
ws.sheet_view.showGridLines = False
title_banner(ws, f"Companies & Organisations  ({len(companies)} records)", 9)
hdr(ws, 2, ["#","Company Name","Type","Country","Reg No","Status","Approval","Credit Limit","Created"])

for i, c in enumerate(companies, 1):
    r = i + 2
    bg = row_bg(i)
    ap_bg = {"approved":"DCFCE7","pending":"FEF3C7","rejected":"FEE2E2"}.get(
        str(c.get("approval_status","")).lower(), "F8FAFC")
    cell(ws, r, 1, i,                              align="center", bg=bg)
    cell(ws, r, 2, fmt(c.get("name")),             bold=True, bg=bg)
    cell(ws, r, 3, fmt(c.get("type")),             bg=bg)
    cell(ws, r, 4, fmt(c.get("country")),          bg=bg)
    cell(ws, r, 5, fmt(c.get("registration_no")),  bg=bg)
    cell(ws, r, 6, fmt(c.get("status")),           bg=bg, align="center")
    cell(ws, r, 7, fmt(c.get("approval_status")),  bg=ap_bg, align="center")
    cell(ws, r, 8, money(c.get("approved_credit_limit")), bg=bg, align="right")
    cell(ws, r, 9, fmt(c.get("created_at",""))[:10], bg=bg, align="center")
    ws.row_dimensions[r].height = 22

widths(ws, [6, 32, 18, 16, 22, 14, 16, 18, 14])
ws.freeze_panes = "A3"

# ── SHEET: Users ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Users")
ws2.sheet_view.showGridLines = False
title_banner(ws2, f"Users / Profiles  ({len(profiles)} records)", 7)
hdr(ws2, 2, ["#","Full Name","Email","Role","Company","Status","Created"])

for i, p in enumerate(profiles, 1):
    r = i + 2
    bg = row_bg(i)
    role_bg = {"admin":"FEF3C7","customer":"DBEAFE","service_provider":"DCFCE7"}.get(
        str(p.get("role","")).lower(), "F8FAFC")
    cell(ws2, r, 1, i,                                      align="center", bg=bg)
    cell(ws2, r, 2, fmt(p.get("full_name")),                bold=True, bg=bg)
    cell(ws2, r, 3, fmt(p.get("email")),                    bg=bg)
    cell(ws2, r, 4, fmt(p.get("role")),                     bg=role_bg, align="center")
    co_id = p.get("company_id")
    cell(ws2, r, 5, co_lookup.get(co_id, fmt(co_id)),       bg=bg)
    cell(ws2, r, 6, fmt(p.get("status","active")),          bg=bg, align="center")
    cell(ws2, r, 7, fmt(p.get("created_at",""))[:10],       bg=bg, align="center")
    ws2.row_dimensions[r].height = 22

widths(ws2, [6, 28, 36, 20, 30, 14, 14])
ws2.freeze_panes = "A3"

# ── SHEET: Secured Jobs ───────────────────────────────────────────────────────
ws3 = wb.create_sheet("Secured Jobs")
ws3.sheet_view.showGridLines = False
title_banner(ws3, f"Secured Jobs  ({len(jobs)} records)", 10)
hdr(ws3, 2, ["#","Job Reference","Customer","Provider","Category","Value","Currency","Status","Payment Model","Created"])

for i, j in enumerate(jobs, 1):
    r = i + 2
    bg = row_bg(i)
    st_bg = {"completed":"DCFCE7","in_progress":"DBEAFE","cancelled":"FEE2E2"}.get(
        str(j.get("job_status","")).lower(), "FEF3C7")
    cell(ws3, r, 1, i,                                           align="center", bg=bg)
    cell(ws3, r, 2, fmt(j.get("job_reference")),                 bold=True, bg=bg, color="0369A1")
    cell(ws3, r, 3, co_lookup.get(j.get("customer_company_id"), fmt(j.get("customer_company_id"))), bg=bg)
    cell(ws3, r, 4, co_lookup.get(j.get("provider_company_id"), fmt(j.get("provider_company_id"))), bg=bg)
    cell(ws3, r, 5, fmt(j.get("service_category")),              bg=bg)
    cell(ws3, r, 6, money(j.get("job_value")),                   bg=bg, align="right")
    cell(ws3, r, 7, fmt(j.get("currency","MYR")),                bg=bg, align="center")
    cell(ws3, r, 8, fmt(j.get("job_status")),                    bg=st_bg, align="center")
    cell(ws3, r, 9, fmt(j.get("payment_model")),                 bg=bg)
    cell(ws3, r, 10, fmt(j.get("created_at",""))[:10],           bg=bg, align="center")
    ws3.row_dimensions[r].height = 22

widths(ws3, [6, 28, 26, 26, 22, 16, 10, 18, 24, 14])
ws3.freeze_panes = "A3"

# ── SHEET: Vendor Credit Terms ────────────────────────────────────────────────
ws4 = wb.create_sheet("Vendor Credit")
ws4.sheet_view.showGridLines = False
title_banner(ws4, f"Vendor Credit Terms — Supplier Obligations  ({len(vendor_terms)} records)", 10)
hdr(ws4, 2, ["#","Reference","Buyer Company","Supplier Name","Invoice Amount","Currency","Due Date","Days Overdue","Payment Status","Created"])

for i, v in enumerate(vendor_terms, 1):
    r = i + 2
    bg = row_bg(i)
    ps = str(v.get("payment_status","")).lower()
    ps_bg = {"paid on time":"DCFCE7","paid late":"FEF3C7","overdue":"FEE2E2",
             "disputed":"F3E8FF","pending":"DBEAFE"}.get(ps, "F8FAFC")
    cell(ws4, r, 1,  i,                                               align="center", bg=bg)
    cell(ws4, r, 2,  fmt(v.get("term_reference")),                    bold=True, bg=bg, color="0369A1")
    cell(ws4, r, 3,  co_lookup.get(v.get("buyer_company_id"), "—"),  bg=bg)
    cell(ws4, r, 4,  fmt(v.get("supplier_name")),                     bold=True, bg=bg)
    cell(ws4, r, 5,  money(v.get("invoice_amount")),                  bg=bg, align="right")
    cell(ws4, r, 6,  fmt(v.get("currency","MYR")),                    bg=bg, align="center")
    cell(ws4, r, 7,  fmt(v.get("due_date",""))[:10],                  bg=bg, align="center")
    cell(ws4, r, 8,  fmt(v.get("days_overdue")),                      bg=("FEE2E2" if v.get("days_overdue",0)>0 else bg), align="center")
    cell(ws4, r, 9,  fmt(v.get("payment_status")),                    bg=ps_bg, align="center")
    cell(ws4, r, 10, fmt(v.get("created_at",""))[:10],                bg=bg, align="center")
    ws4.row_dimensions[r].height = 22

widths(ws4, [6, 26, 28, 28, 18, 10, 14, 14, 20, 14])
ws4.freeze_panes = "A3"

# ── SHEET: TradeFlow Requests ─────────────────────────────────────────────────
ws5 = wb.create_sheet("TradeFlow Requests")
ws5.sheet_view.showGridLines = False
title_banner(ws5, f"TradeFlow Requests  ({len(tradeflows)} records)", 9)
hdr(ws5, 2, ["#","Reference","Customer","Supplier Name","Trade Amount","Currency","Payment Model","Status","Created"])

for i, t in enumerate(tradeflows, 1):
    r = i + 2
    bg = row_bg(i)
    st = str(t.get("status","")).lower()
    st_bg = {"approved":"DCFCE7","pending":"FEF3C7","rejected":"FEE2E2",
             "disbursed":"DBEAFE","completed":"DCFCE7"}.get(st, "F8FAFC")
    cell(ws5, r, 1, i,                                               align="center", bg=bg)
    cell(ws5, r, 2, fmt(t.get("tradeflow_reference")),               bold=True, bg=bg, color="0369A1")
    cell(ws5, r, 3, co_lookup.get(t.get("customer_id"), "—"),       bg=bg)
    cell(ws5, r, 4, fmt(t.get("supplier_name")),                     bold=True, bg=bg)
    cell(ws5, r, 5, money(t.get("trade_amount")),                    bg=bg, align="right")
    cell(ws5, r, 6, fmt(t.get("currency","MYR")),                    bg=bg, align="center")
    cell(ws5, r, 7, fmt(t.get("payment_model")),                     bg=bg)
    cell(ws5, r, 8, fmt(t.get("status")),                            bg=st_bg, align="center")
    cell(ws5, r, 9, fmt(t.get("created_at",""))[:10],                bg=bg, align="center")
    ws5.row_dimensions[r].height = 22

widths(ws5, [6, 26, 28, 28, 18, 10, 24, 18, 14])
ws5.freeze_panes = "A3"

# ── SHEET: RFQs ──────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("RFQ Requests")
ws6.sheet_view.showGridLines = False
title_banner(ws6, f"RFQ Requests  ({len(rfqs)} records)", 9)
hdr(ws6, 2, ["#","RFQ Reference","Customer","Category","Origin","Destination","Status","Quotes","Created"])

for i, r_ in enumerate(rfqs, 1):
    r = i + 2
    bg = row_bg(i)
    rfq_quotes_count = sum(1 for q in quotes if q.get("rfq_id") == r_.get("id"))
    cell(ws6, r, 1, i,                                                    align="center", bg=bg)
    cell(ws6, r, 2, fmt(r_.get("rfq_reference")),                         bold=True, bg=bg, color="0369A1")
    cell(ws6, r, 3, co_lookup.get(r_.get("customer_company_id"), "—"),   bg=bg)
    cell(ws6, r, 4, fmt(r_.get("service_category")),                      bg=bg)
    cell(ws6, r, 5, fmt(r_.get("origin") or r_.get("origin_port","")),   bg=bg)
    cell(ws6, r, 6, fmt(r_.get("destination") or r_.get("destination_port","")), bg=bg)
    cell(ws6, r, 7, fmt(r_.get("rfq_status")),                            bg=bg, align="center")
    cell(ws6, r, 8, rfq_quotes_count,                                     bg="DBEAFE" if rfq_quotes_count else bg, align="center")
    cell(ws6, r, 9, fmt(r_.get("created_at",""))[:10],                    bg=bg, align="center")
    ws6.row_dimensions[r].height = 22

widths(ws6, [6, 26, 28, 22, 22, 22, 18, 12, 14])
ws6.freeze_panes = "A3"

# ── SHEET: Service Listings ───────────────────────────────────────────────────
ws7 = wb.create_sheet("Service Listings")
ws7.sheet_view.showGridLines = False
title_banner(ws7, f"LSP Service Listings  ({len(listings)} records)", 9)
hdr(ws7, 2, ["#","Reference","Provider","Category","Service Name","Origin","Destination","Base Rate","Approval"])

for i, l in enumerate(listings, 1):
    r = i + 2
    bg = row_bg(i)
    ap = str(l.get("approval_status","")).lower()
    ap_bg = {"approved":"DCFCE7","pending":"FEF3C7","rejected":"FEE2E2"}.get(ap, "F8FAFC")
    cell(ws7, r, 1, i,                                                              align="center", bg=bg)
    cell(ws7, r, 2, fmt(l.get("listing_reference")),                                bold=True, bg=bg, color="0369A1")
    cell(ws7, r, 3, co_lookup.get(l.get("provider_company_id"), "—"),              bg=bg)
    cell(ws7, r, 4, fmt(l.get("service_category")),                                 bg=bg)
    cell(ws7, r, 5, fmt(l.get("service_name")),                                     bold=True, bg=bg)
    cell(ws7, r, 6, fmt(l.get("origin_country") or l.get("origin_port","")),       bg=bg)
    cell(ws7, r, 7, fmt(l.get("destination_country") or l.get("destination_port","")), bg=bg)
    cell(ws7, r, 8, money(l.get("base_rate")),                                      bg=bg, align="right")
    cell(ws7, r, 9, fmt(l.get("approval_status")),                                  bg=ap_bg, align="center")
    ws7.row_dimensions[r].height = 22

widths(ws7, [6, 26, 28, 22, 30, 20, 20, 16, 18])
ws7.freeze_panes = "A3"

# ── SHEET: Shipment Bundles ───────────────────────────────────────────────────
ws8 = wb.create_sheet("Shipment Bundles")
ws8.sheet_view.showGridLines = False
title_banner(ws8, f"Shipment Bundles  ({len(bundles)} records)", 9)
hdr(ws8, 2, ["#","Bundle Reference","Customer","Total Value","Currency","Origin","Destination","Status","Created"])

for i, b in enumerate(bundles, 1):
    r = i + 2
    bg = row_bg(i)
    st_bg = {"completed":"DCFCE7","active":"DBEAFE","cancelled":"FEE2E2"}.get(
        str(b.get("bundle_status","")).lower(), "FEF3C7")
    cell(ws8, r, 1, i,                                              align="center", bg=bg)
    cell(ws8, r, 2, fmt(b.get("bundle_reference")),                 bold=True, bg=bg, color="0369A1")
    cell(ws8, r, 3, co_lookup.get(b.get("customer_company_id"), "—"), bg=bg)
    cell(ws8, r, 4, money(b.get("total_value")),                    bg=bg, align="right")
    cell(ws8, r, 5, fmt(b.get("currency","MYR")),                   bg=bg, align="center")
    cell(ws8, r, 6, fmt(b.get("origin")),                           bg=bg)
    cell(ws8, r, 7, fmt(b.get("destination")),                      bg=bg)
    cell(ws8, r, 8, fmt(b.get("bundle_status")),                    bg=st_bg, align="center")
    cell(ws8, r, 9, fmt(b.get("created_at",""))[:10],               bg=bg, align="center")
    ws8.row_dimensions[r].height = 22

widths(ws8, [6, 28, 28, 16, 10, 22, 22, 18, 14])
ws8.freeze_panes = "A3"

# ── SHEET: TradeCycle Wallets ─────────────────────────────────────────────────
ws9 = wb.create_sheet("Wallets")
ws9.sheet_view.showGridLines = False
title_banner(ws9, f"TradeCycle Wallets  ({len(wallets)} records)", 8)
hdr(ws9, 2, ["#","Company","Currency","Total Balance","Available","Reserved","Settled","Status"])

for i, w in enumerate(wallets, 1):
    r = i + 2
    bg = row_bg(i)
    cell(ws9, r, 1, i,                                              align="center", bg=bg)
    cell(ws9, r, 2, co_lookup.get(w.get("company_id"), "—"),       bold=True, bg=bg)
    cell(ws9, r, 3, fmt(w.get("currency","MYR")),                   bg=bg, align="center")
    cell(ws9, r, 4, money(w.get("total_balance")),                  bg=bg, align="right", bold=True)
    cell(ws9, r, 5, money(w.get("available_balance")),              bg="DCFCE7", align="right")
    cell(ws9, r, 6, money(w.get("reserved_balance")),               bg="FEF3C7", align="right")
    cell(ws9, r, 7, money(w.get("settled_balance")),                bg=bg, align="right")
    cell(ws9, r, 8, fmt(w.get("wallet_status","active")),           bg=bg, align="center")
    ws9.row_dimensions[r].height = 22

widths(ws9, [6, 32, 12, 18, 18, 18, 18, 14])
ws9.freeze_panes = "A3"

# ── SHEET: Wallet Transactions ────────────────────────────────────────────────
ws10 = wb.create_sheet("Wallet Transactions")
ws10.sheet_view.showGridLines = False
title_banner(ws10, f"Wallet Transactions  ({len(wallet_txns)} records)", 8)
hdr(ws10, 2, ["#","Company","Txn Type","Amount","Currency","Direction","Reference","Created"])

for i, t in enumerate(wallet_txns, 1):
    r = i + 2
    bg = row_bg(i)
    direction = str(t.get("direction","")).lower()
    dir_bg = "DCFCE7" if direction == "credit" else ("FEE2E2" if direction == "debit" else bg)
    cell(ws10, r, 1, i,                                               align="center", bg=bg)
    cell(ws10, r, 2, co_lookup.get(t.get("company_id"), "—"),        bg=bg)
    cell(ws10, r, 3, fmt(t.get("transaction_type")),                  bg=bg)
    cell(ws10, r, 4, money(t.get("amount")),                          bg=bg, align="right", bold=True)
    cell(ws10, r, 5, fmt(t.get("currency","MYR")),                    bg=bg, align="center")
    cell(ws10, r, 6, fmt(t.get("direction")),                         bg=dir_bg, align="center")
    cell(ws10, r, 7, fmt(t.get("reference_id")),                      bg=bg)
    cell(ws10, r, 8, fmt(t.get("created_at",""))[:10],                bg=bg, align="center")
    ws10.row_dimensions[r].height = 22

widths(ws10, [6, 30, 24, 16, 10, 14, 28, 14])
ws10.freeze_panes = "A3"

# ── SHEET: Intelligence Scores ────────────────────────────────────────────────
ws11 = wb.create_sheet("Intelligence Scores")
ws11.sheet_view.showGridLines = False
title_banner(ws11, f"Company Intelligence Scores  ({len(scores)} records)", 10)
hdr(ws11, 2, ["#","Company","Overall","Payment","Document","Shipment","Risk Level","Financing Readiness","Rec. Limit (MYR)","Calculated"])

for i, s in enumerate(scores, 1):
    r = i + 2
    bg = row_bg(i)
    overall = s.get("overall_score") or 0
    score_bg = "DCFCE7" if overall >= 75 else ("FEF3C7" if overall >= 50 else "FEE2E2")
    rl_bg = {"Low":"DCFCE7","Medium":"FEF3C7","High":"FEE2E2","Critical":"F3E8FF"}.get(
        str(s.get("risk_level","")), "F8FAFC")
    cell(ws11, r, 1,  i,                                                align="center", bg=bg)
    cell(ws11, r, 2,  co_lookup.get(s.get("company_id"), "—"),         bold=True, bg=bg)
    cell(ws11, r, 3,  money(s.get("overall_score")),                    bg=score_bg, align="center", bold=True)
    cell(ws11, r, 4,  money(s.get("payment_behaviour_score")),          bg=bg, align="center")
    cell(ws11, r, 5,  money(s.get("document_accuracy_score")),          bg=bg, align="center")
    cell(ws11, r, 6,  money(s.get("shipment_performance_score")),       bg=bg, align="center")
    cell(ws11, r, 7,  fmt(s.get("risk_level")),                         bg=rl_bg, align="center")
    cell(ws11, r, 8,  fmt(s.get("financing_readiness")),                bg=bg, align="center")
    cell(ws11, r, 9,  money(s.get("recommended_limit")),                bg=bg, align="right")
    cell(ws11, r, 10, fmt(s.get("calculated_at",""))[:10],              bg=bg, align="center")
    ws11.row_dimensions[r].height = 22

widths(ws11, [6, 30, 12, 12, 12, 12, 14, 26, 20, 14])
ws11.freeze_panes = "A3"

# ── SHEET: Risk Signals ───────────────────────────────────────────────────────
ws12 = wb.create_sheet("Risk Signals")
ws12.sheet_view.showGridLines = False
title_banner(ws12, f"Intelligence Risk Signals  ({len(risk_signals)} records)", 8)
hdr(ws12, 2, ["#","Signal Reference","Company","Signal Type","Severity","Status","Description","Created"])

for i, s in enumerate(risk_signals, 1):
    r = i + 2
    bg = row_bg(i)
    sev_bg = {"Critical":"FEE2E2","High":"FEF3C7","Medium":"FEF9C3","Low":"DBEAFE"}.get(
        str(s.get("severity","")), bg)
    cell(ws12, r, 1, i,                                                  align="center", bg=bg)
    cell(ws12, r, 2, fmt(s.get("signal_reference")),                     bold=True, bg=bg, color="0369A1")
    cell(ws12, r, 3, co_lookup.get(s.get("related_company_id"), "—"),   bg=bg)
    cell(ws12, r, 4, fmt(s.get("signal_type")),                          bg=bg)
    cell(ws12, r, 5, fmt(s.get("severity")),                             bg=sev_bg, align="center")
    cell(ws12, r, 6, fmt(s.get("status")),                               bg=bg, align="center")
    cell(ws12, r, 7, fmt(s.get("description"))[:100],                    bg=bg, wrap=True)
    cell(ws12, r, 8, fmt(s.get("created_at",""))[:10],                   bg=bg, align="center")
    ws12.row_dimensions[r].height = 30

widths(ws12, [6, 26, 28, 22, 14, 14, 50, 14])
ws12.freeze_panes = "A3"

# ── SHEET: Trade Chains ───────────────────────────────────────────────────────
ws13 = wb.create_sheet("Trade Chains")
ws13.sheet_view.showGridLines = False
title_banner(ws13, f"Trade Chains  ({len(chains)} records)", 8)
hdr(ws13, 2, ["#","Chain Reference","Initiating Company","Trade Type","Status","Total Value","Currency","Created"])

for i, c in enumerate(chains, 1):
    r = i + 2
    bg = row_bg(i)
    cell(ws13, r, 1, i,                                                     align="center", bg=bg)
    cell(ws13, r, 2, fmt(c.get("chain_reference")),                          bold=True, bg=bg, color="0369A1")
    cell(ws13, r, 3, co_lookup.get(c.get("initiating_company_id"), "—"),   bg=bg)
    cell(ws13, r, 4, fmt(c.get("trade_type")),                               bg=bg)
    cell(ws13, r, 5, fmt(c.get("chain_status")),                             bg=bg, align="center")
    cell(ws13, r, 6, money(c.get("total_value")),                            bg=bg, align="right")
    cell(ws13, r, 7, fmt(c.get("currency","MYR")),                           bg=bg, align="center")
    cell(ws13, r, 8, fmt(c.get("created_at",""))[:10],                       bg=bg, align="center")
    ws13.row_dimensions[r].height = 22

widths(ws13, [6, 28, 30, 22, 18, 18, 10, 14])
ws13.freeze_panes = "A3"

# ── SHEET: Summary ────────────────────────────────────────────────────────────
wss = wb.create_sheet("Summary", 0)
wss.sheet_view.showGridLines = False
wss.merge_cells("A1:C1")
c = wss.cell(row=1, column=1, value="Nexum SecureFlow — Live Data Export")
c.font = Font(name="Arial", bold=True, color="FFFFFF", size=16)
c.fill = PatternFill("solid", fgColor="0B1929")
c.alignment = Alignment(horizontal="center", vertical="center")
wss.row_dimensions[1].height = 50

hdr(wss, 2, ["Data Category", "Record Count", "Sheet"], bg="0891B2")

summary_data = [
    ("Companies / Organisations",   len(companies),    "Companies"),
    ("Users / Profiles",             len(profiles),     "Users"),
    ("Secured Jobs",                 len(jobs),         "Secured Jobs"),
    ("Vendor Credit Terms",          len(vendor_terms), "Vendor Credit"),
    ("TradeFlow Requests",           len(tradeflows),   "TradeFlow Requests"),
    ("RFQ Requests",                 len(rfqs),         "RFQ Requests"),
    ("RFQ Quotes",                   len(quotes),       "(merged in RFQ sheet)"),
    ("Service Listings (LSP)",       len(listings),     "Service Listings"),
    ("Shipment Bundles",             len(bundles),      "Shipment Bundles"),
    ("TradeCycle Wallets",           len(wallets),      "Wallets"),
    ("Wallet Transactions",          len(wallet_txns),  "Wallet Transactions"),
    ("Intelligence Scores",          len(scores),       "Intelligence Scores"),
    ("Intelligence Risk Signals",    len(risk_signals), "Risk Signals"),
    ("Trade Chains",                 len(chains),       "Trade Chains"),
]

total = sum(x[1] for x in summary_data)

for i, (label, count, sheet) in enumerate(summary_data):
    r = i + 3
    bg = "F8FAFC" if i % 2 == 0 else "FFFFFF"
    cell(wss, r, 1, label,  bold=True, bg=bg)
    cell(wss, r, 2, count,  bg="DBEAFE" if count > 0 else "FEE2E2", align="center", bold=True, color="0369A1")
    cell(wss, r, 3, sheet,  bg=bg, color="64748B")
    wss.row_dimensions[r].height = 26

# Total row
r_total = len(summary_data) + 3
cell(wss, r_total, 1, "TOTAL RECORDS", bold=True, bg="0369A1", color="FFFFFF", align="center")
cell(wss, r_total, 2, total,           bold=True, bg="0369A1", color="FFFFFF", align="center")
cell(wss, r_total, 3, "",              bg="0369A1")
wss.row_dimensions[r_total].height = 28

widths(wss, [40, 18, 34])

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = Path(__file__).parent / "nexum-live-data-export.xlsx"
wb.save(out_path)

print(f"\n✅  Saved: {out_path}")
print(f"\n📊  Summary:")
for label, count, _ in summary_data:
    status = "✓" if count > 0 else "○"
    print(f"  {status}  {label}: {count}")
print(f"\n  TOTAL: {total} records across {len(summary_data)} tables")
